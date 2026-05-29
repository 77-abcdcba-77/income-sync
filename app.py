from __future__ import annotations

import json
import os
import re
import sqlite3
import sys
import webbrowser
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from tempfile import NamedTemporaryFile
from threading import Timer

from flask import Flask, jsonify, redirect, render_template, request, send_file

from sync_engine import (
    get_device_id,
    get_sync_state,
    init_sync_tables,
    log_change,
    serialize_record,
    start_sync_loop,
    sync_now,
)

def resource_path(relative_path: str) -> Path:
    """兼容源码运行和 PyInstaller 打包后读取 templates/static 等资源。"""
    if hasattr(sys, "_MEIPASS"):
        return Path(sys._MEIPASS) / relative_path
    return Path(__file__).resolve().parent / relative_path


def app_base_dir() -> Path:
    """数据库、导出文件放在源码目录或 exe 同级目录，避免打包后数据丢失。"""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


BASE_DIR = app_base_dir()
DATA_DIR = BASE_DIR / "data"
EXPORT_DIR = BASE_DIR / "exports"
DB_PATH = DATA_DIR / "records.db"
SEED_PATH = DATA_DIR / "seed.json"
BUNDLED_SEED_PATH = resource_path("data/seed.json")

app = Flask(
    __name__,
    template_folder=str(resource_path("templates")),
    static_folder=str(resource_path("static")),
)


def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def today_text() -> str:
    return date.today().strftime("%Y-%m-%d")


def get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def as_float(value) -> float:
    if value in (None, ""):
        return 0.0
    try:
        if isinstance(value, str):
            value = (
                value.replace(",", "")
                .replace("￥", "")
                .replace("¥", "")
                .replace("元", "")
                .strip()
            )
        return round(float(value), 2)
    except (TypeError, ValueError):
        return 0.0


def clean_text(value) -> str:
    return str(value or "").strip()


def clean_date(value) -> str:
    if isinstance(value, datetime):
        return value.date().strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    value = clean_text(value)
    if not value:
        return today_text()
    match = re.match(r"^(\d{4}-\d{1,2}-\d{1,2})", value)
    if not match:
        return today_text()
    parts = match.group(1).split("-")
    try:
        y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
        return date(y, m, d).strftime("%Y-%m-%d")
    except ValueError:
        return today_text()


def normalize_order(payload: dict) -> dict:
    return {
        "wechat": clean_text(payload.get("wechat")),
        "task_name": clean_text(payload.get("task_name")),
        "order_no": clean_text(payload.get("order_no")),
        "deadline_status": clean_text(payload.get("deadline_status")),
        "accepted_date": clean_date(payload.get("accepted_date")),
        "price": as_float(payload.get("price")),
    }


def normalize_payment(payload: dict) -> dict:
    return {
        "pay_date": clean_date(payload.get("pay_date")),
        "amount": as_float(payload.get("amount")),
        "note": clean_text(payload.get("note")),
    }


def normalize_adjustment(payload: dict) -> dict:
    return {
        "change_date": clean_date(payload.get("change_date")),
        "amount": as_float(payload.get("amount")),
        "note": clean_text(payload.get("note")),
    }


def normalize_expense(payload: dict) -> dict:
    return {
        "expense_date": clean_date(payload.get("expense_date")),
        "name": clean_text(payload.get("name")),
        "amount": as_float(payload.get("amount")),
        "note": clean_text(payload.get("note")),
    }


def table_columns(conn: sqlite3.Connection, table_name: str) -> set[str]:
    rows = conn.execute(f"PRAGMA table_info({table_name})").fetchall()
    return {row[1] for row in rows}


def add_column_if_missing(conn: sqlite3.Connection, table_name: str, column_name: str, ddl: str) -> None:
    if column_name not in table_columns(conn, table_name):
        conn.execute(f"ALTER TABLE {table_name} ADD COLUMN {ddl}")


def init_db() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    with get_conn() as conn:
        # records 保留原表名，避免旧版本数据丢失；现在它表示“订单主表”。
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                wechat TEXT,
                task_name TEXT,
                order_no TEXT,
                deadline_status TEXT,
                accepted_date TEXT NOT NULL DEFAULT '',
                price REAL NOT NULL DEFAULT 0,
                paid REAL NOT NULL DEFAULT 0,
                remaining REAL NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
        add_column_if_missing(conn, "records", "accepted_date", "accepted_date TEXT NOT NULL DEFAULT ''")
        add_column_if_missing(conn, "records", "updated_at", "updated_at TEXT NOT NULL DEFAULT ''")

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS payments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                record_id INTEGER NOT NULL,
                pay_date TEXT NOT NULL,
                amount REAL NOT NULL DEFAULT 0,
                note TEXT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(record_id) REFERENCES records(id) ON DELETE CASCADE
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS budget_changes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                record_id INTEGER NOT NULL,
                change_date TEXT NOT NULL,
                amount REAL NOT NULL DEFAULT 0,
                note TEXT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(record_id) REFERENCES records(id) ON DELETE CASCADE
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS expenses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                expense_date TEXT NOT NULL,
                name TEXT NOT NULL,
                amount REAL NOT NULL DEFAULT 0,
                note TEXT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
            """
        )

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS history_versions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                created_at TEXT NOT NULL,
                label TEXT NOT NULL,
                reason TEXT,
                source_file TEXT,
                record_count INTEGER NOT NULL DEFAULT 0,
                payment_count INTEGER NOT NULL DEFAULT 0,
                adjustment_count INTEGER NOT NULL DEFAULT 0,
                expense_count INTEGER NOT NULL DEFAULT 0,
                summary_json TEXT,
                payload_json TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS import_batches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                version_id INTEGER,
                imported_at TEXT NOT NULL,
                source_file TEXT,
                sheet_name TEXT,
                parsed_count INTEGER NOT NULL DEFAULT 0,
                created_count INTEGER NOT NULL DEFAULT 0,
                updated_count INTEGER NOT NULL DEFAULT 0,
                unchanged_count INTEGER NOT NULL DEFAULT 0,
                skipped_count INTEGER NOT NULL DEFAULT 0,
                payment_adjusted_count INTEGER NOT NULL DEFAULT 0,
                summary_json TEXT,
                FOREIGN KEY(version_id) REFERENCES history_versions(id) ON DELETE SET NULL
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS idx_records_order_no ON records(order_no)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_records_client_task ON records(wechat, task_name)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_payments_record_id ON payments(record_id)")

        # 同步所需的 change_log 表
        init_sync_tables(conn)

        # 旧数据没有接单日期时，补为创建日期；仍为空则补为今天。
        conn.execute(
            """
            UPDATE records
            SET accepted_date = COALESCE(NULLIF(substr(created_at, 1, 10), ''), ?)
            WHERE accepted_date IS NULL OR accepted_date = ''
            """,
            (today_text(),),
        )
        conn.execute(
            """
            UPDATE records
            SET updated_at = COALESCE(NULLIF(updated_at, ''), COALESCE(created_at, ?))
            WHERE updated_at IS NULL OR updated_at = ''
            """,
            (now_text(),),
        )

        count = conn.execute("SELECT COUNT(*) FROM records").fetchone()[0]
        seed_source = SEED_PATH if SEED_PATH.exists() else BUNDLED_SEED_PATH
        if count == 0 and seed_source.exists():
            seed = json.loads(seed_source.read_text(encoding="utf-8"))
            for item in seed:
                order = normalize_order({**item, "accepted_date": item.get("accepted_date") or today_text()})
                cur = conn.execute(
                    """
                    INSERT INTO records
                    (wechat, task_name, order_no, deadline_status, accepted_date, price, paid, remaining, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        order["wechat"], order["task_name"], order["order_no"], order["deadline_status"],
                        order["accepted_date"], order["price"], as_float(item.get("paid")),
                        as_float(item.get("remaining")), now_text(), now_text(),
                    ),
                )
                paid = as_float(item.get("paid"))
                if paid:
                    conn.execute(
                        """
                        INSERT INTO payments (record_id, pay_date, amount, note, created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?)
                        """,
                        (cur.lastrowid, order["accepted_date"], paid, "历史到账导入", now_text(), now_text()),
                    )

        # 兼容从旧版 records.paid 迁移：当 payments 为空时，把旧到账拆成一条历史付款流水。
        payment_count = conn.execute("SELECT COUNT(*) FROM payments").fetchone()[0]
        if payment_count == 0 and "paid" in table_columns(conn, "records"):
            old_rows = conn.execute("SELECT id, paid, accepted_date FROM records WHERE COALESCE(paid, 0) != 0").fetchall()
            for row in old_rows:
                conn.execute(
                    """
                    INSERT INTO payments (record_id, pay_date, amount, note, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (row["id"], clean_date(row["accepted_date"]), as_float(row["paid"]), "历史到账导入", now_text(), now_text()),
                )
        conn.commit()


def status_bucket(value: str) -> str:
    v = clean_text(value)
    if not v:
        return "未填写"
    if v in ("已交", "修改", "进行中", "待交"):
        return v
    if re.match(r"^\d{4}-\d{2}-\d{2}", v):
        return "截止日期"
    return "其他"


def fetch_records() -> list[dict]:
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT
                r.id, r.wechat, r.task_name, r.order_no, r.deadline_status,
                r.accepted_date, r.price, r.created_at, r.updated_at,
                COALESCE(p.total_paid, 0) AS paid,
                COALESCE(b.total_adjustment, 0) AS adjustment
            FROM records r
            LEFT JOIN (
                SELECT record_id, SUM(amount) AS total_paid
                FROM payments
                GROUP BY record_id
            ) p ON p.record_id = r.id
            LEFT JOIN (
                SELECT record_id, SUM(amount) AS total_adjustment
                FROM budget_changes
                GROUP BY record_id
            ) b ON b.record_id = r.id
            ORDER BY r.accepted_date DESC, r.id DESC
            """
        ).fetchall()

    result = []
    for row in rows:
        price = as_float(row["price"])
        adjustment = as_float(row["adjustment"])
        total_price = round(price + adjustment, 2)
        paid = as_float(row["paid"])
        result.append(
            {
                "id": row["id"],
                "wechat": row["wechat"] or "",
                "task_name": row["task_name"] or "",
                "order_no": row["order_no"] or "",
                "deadline_status": row["deadline_status"] or "",
                "accepted_date": row["accepted_date"] or "",
                "price": price,
                "adjustment": adjustment,
                "total_price": total_price,
                "paid": paid,
                "remaining": round(total_price - paid, 2),
                "created_at": row["created_at"] or "",
                "updated_at": row["updated_at"] or "",
            }
        )
    return result


def fetch_payments(record_id: int | None = None) -> list[dict]:
    where = ""
    params: tuple = ()
    if record_id is not None:
        where = "WHERE p.record_id = ?"
        params = (record_id,)
    with get_conn() as conn:
        rows = conn.execute(
            f"""
            SELECT p.*, r.wechat, r.task_name, r.order_no
            FROM payments p
            LEFT JOIN records r ON r.id = p.record_id
            {where}
            ORDER BY p.pay_date DESC, p.id DESC
            """,
            params,
        ).fetchall()
    return [
        {
            "id": row["id"],
            "record_id": row["record_id"],
            "pay_date": row["pay_date"],
            "amount": as_float(row["amount"]),
            "note": row["note"] or "",
            "wechat": row["wechat"] or "",
            "task_name": row["task_name"] or "",
            "order_no": row["order_no"] or "",
        }
        for row in rows
    ]


def fetch_adjustments(record_id: int | None = None) -> list[dict]:
    where = ""
    params: tuple = ()
    if record_id is not None:
        where = "WHERE b.record_id = ?"
        params = (record_id,)
    with get_conn() as conn:
        rows = conn.execute(
            f"""
            SELECT b.*, r.wechat, r.task_name, r.order_no
            FROM budget_changes b
            LEFT JOIN records r ON r.id = b.record_id
            {where}
            ORDER BY b.change_date DESC, b.id DESC
            """,
            params,
        ).fetchall()
    return [
        {
            "id": row["id"],
            "record_id": row["record_id"],
            "change_date": row["change_date"],
            "amount": as_float(row["amount"]),
            "note": row["note"] or "",
            "wechat": row["wechat"] or "",
            "task_name": row["task_name"] or "",
            "order_no": row["order_no"] or "",
        }
        for row in rows
    ]


def fetch_expenses() -> list[dict]:
    with get_conn() as conn:
        rows = conn.execute("SELECT * FROM expenses ORDER BY expense_date DESC, id DESC").fetchall()
    return [
        {
            "id": row["id"],
            "expense_date": row["expense_date"],
            "name": row["name"] or "",
            "amount": as_float(row["amount"]),
            "note": row["note"] or "",
            "created_at": row["created_at"] or "",
            "updated_at": row["updated_at"] or "",
        }
        for row in rows
    ]


def table_rows(conn: sqlite3.Connection, table_name: str) -> list[dict]:
    rows = conn.execute(f"SELECT * FROM {table_name} ORDER BY id").fetchall()
    return [dict(row) for row in rows]


def build_snapshot_payload(conn: sqlite3.Connection) -> dict:
    return {
        "records": table_rows(conn, "records"),
        "payments": table_rows(conn, "payments"),
        "budget_changes": table_rows(conn, "budget_changes"),
        "expenses": table_rows(conn, "expenses"),
    }


def create_history_version(
    conn: sqlite3.Connection,
    label: str,
    reason: str = "",
    source_file: str = "",
    summary: dict | None = None,
) -> int:
    payload = build_snapshot_payload(conn)
    summary = summary or {}
    cur = conn.execute(
        """
        INSERT INTO history_versions
        (created_at, label, reason, source_file, record_count, payment_count,
         adjustment_count, expense_count, summary_json, payload_json)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            now_text(),
            label,
            reason,
            source_file,
            len(payload["records"]),
            len(payload["payments"]),
            len(payload["budget_changes"]),
            len(payload["expenses"]),
            json.dumps(summary, ensure_ascii=False),
            json.dumps(payload, ensure_ascii=False),
        ),
    )
    return int(cur.lastrowid)


def fetch_history_versions(limit: int = 30) -> list[dict]:
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT id, created_at, label, reason, source_file, record_count,
                   payment_count, adjustment_count, expense_count, summary_json
            FROM history_versions
            ORDER BY id DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()
    result = []
    for row in rows:
        summary = {}
        if row["summary_json"]:
            try:
                summary = json.loads(row["summary_json"])
            except json.JSONDecodeError:
                summary = {}
        result.append(
            {
                "id": row["id"],
                "created_at": row["created_at"],
                "label": row["label"],
                "reason": row["reason"] or "",
                "source_file": row["source_file"] or "",
                "record_count": row["record_count"],
                "payment_count": row["payment_count"],
                "adjustment_count": row["adjustment_count"],
                "expense_count": row["expense_count"],
                "summary": summary,
            }
        )
    return result


def insert_snapshot_rows(conn: sqlite3.Connection, table_name: str, rows: list[dict]) -> None:
    if not rows:
        return
    valid_columns = table_columns(conn, table_name)
    for row in rows:
        data = {key: value for key, value in row.items() if key in valid_columns}
        columns = list(data.keys())
        placeholders = ", ".join("?" for _ in columns)
        conn.execute(
            f"INSERT INTO {table_name} ({', '.join(columns)}) VALUES ({placeholders})",
            tuple(data[column] for column in columns),
        )


def restore_history_version(version_id: int) -> int:
    with get_conn() as conn:
        row = conn.execute("SELECT * FROM history_versions WHERE id = ?", (version_id,)).fetchone()
        if not row:
            raise ValueError("历史版本不存在")
        payload = json.loads(row["payload_json"])
        backup_id = create_history_version(
            conn,
            f"回退前自动备份 {now_text()}",
            reason=f"准备回退到版本 #{version_id}",
            source_file=row["source_file"] or "",
            summary={"rollback_target": version_id},
        )
        conn.execute("DELETE FROM payments")
        conn.execute("DELETE FROM budget_changes")
        conn.execute("DELETE FROM expenses")
        conn.execute("DELETE FROM records")
        insert_snapshot_rows(conn, "records", payload.get("records", []))
        insert_snapshot_rows(conn, "payments", payload.get("payments", []))
        insert_snapshot_rows(conn, "budget_changes", payload.get("budget_changes", []))
        insert_snapshot_rows(conn, "expenses", payload.get("expenses", []))
        conn.commit()
    return backup_id


HEADER_ALIASES = {
    "wechat": ("微信号", "微信", "客户微信", "客户", "客户号"),
    "task_name": ("任务名字", "任务名称", "任务", "项目名称", "项目", "名称"),
    "order_no": ("单号", "订单号", "订单编号", "编号", "订单"),
    "deadline_status": ("截止时间", "截止日期", "截止/状态", "状态", "进度"),
    "accepted_date": ("接单日期", "接单时间", "日期", "创建日期"),
    "price": ("价格", "基础价格", "订单总价", "总价格", "总价", "金额"),
    "paid": ("到账", "已到账", "已付", "已收", "收款", "付款"),
    "remaining": ("剩余", "剩余未收", "未收", "尾款", "待收"),
}


def normalized_header(value) -> str:
    text = clean_text(value).lower()
    return re.sub(r"[\s/\\:_：\-（）()【】\[\]·.]+", "", text)


HEADER_LOOKUP = {
    normalized_header(alias): field
    for field, aliases in HEADER_ALIASES.items()
    for alias in aliases
}


def parse_import_workbook(file_obj) -> dict:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError(f"缺少 openpyxl：{exc}，请先安装 requirements.txt") from exc

    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    candidates = []
    for ws in wb.worksheets:
        for row_index, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True), start=1):
            mapping = {}
            for col_index, value in enumerate(row, start=1):
                field = HEADER_LOOKUP.get(normalized_header(value))
                if field and field not in mapping:
                    mapping[field] = col_index - 1
            score = len(set(mapping) & {"wechat", "task_name", "order_no", "price", "paid", "remaining"})
            if score >= 4:
                priority = 2 if ws.title == "明细表" else 0
                candidates.append((score + priority, ws.title, row_index, mapping))
                break
    if not candidates:
        raise ValueError("没有识别到明细表表头，请确认包含“微信号、任务名字、单号、价格、到账”等列")

    _, sheet_name, header_row, mapping = sorted(candidates, reverse=True)[0]
    ws = wb[sheet_name]
    rows = []
    skipped = []
    for row_number, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        values = {field: row[col] if col < len(row) else None for field, col in mapping.items()}
        first_text = clean_text(row[0] if row else "")
        if not any(clean_text(v) for v in values.values()):
            continue
        if first_text.startswith("合计"):
            continue
        item = {
            "source_row": row_number,
            "wechat": clean_text(values.get("wechat")),
            "task_name": clean_text(values.get("task_name")),
            "order_no": clean_text(values.get("order_no")),
            "deadline_status": clean_text(values.get("deadline_status")),
            "accepted_date": clean_date(values.get("accepted_date")) if values.get("accepted_date") else "",
            "price": as_float(values.get("price")),
            "paid": as_float(values.get("paid")),
            "remaining": as_float(values.get("remaining")),
        }
        if not item["paid"] and item["price"] and item["remaining"]:
            item["paid"] = round(item["price"] - item["remaining"], 2)
        if not (item["wechat"] or item["task_name"] or item["order_no"]):
            skipped.append({"row": row_number, "reason": "缺少客户、任务和单号"})
            continue
        rows.append(item)
    return {
        "sheet_name": sheet_name,
        "header_row": header_row,
        "columns": sorted(mapping.keys()),
        "rows": rows,
        "skipped": skipped,
    }


def existing_records_for_import(conn: sqlite3.Connection) -> list[dict]:
    rows = conn.execute(
        """
        SELECT
            r.id, r.wechat, r.task_name, r.order_no, r.deadline_status,
            r.accepted_date, r.price, COALESCE(p.total_paid, 0) AS paid
        FROM records r
        LEFT JOIN (
            SELECT record_id, SUM(amount) AS total_paid
            FROM payments
            GROUP BY record_id
        ) p ON p.record_id = r.id
        """
    ).fetchall()
    return [dict(row) for row in rows]


def diff_order_fields(existing: dict, imported: dict) -> list[str]:
    fields = ["wechat", "task_name", "order_no", "deadline_status", "price"]
    if imported.get("accepted_date"):
        fields.append("accepted_date")
    changed = []
    for field in fields:
        old_value = as_float(existing.get(field)) if field == "price" else clean_text(existing.get(field))
        new_value = as_float(imported.get(field)) if field == "price" else clean_text(imported.get(field))
        if field != "price" and not new_value:
            continue
        if old_value != new_value:
            changed.append(field)
    return changed


def build_import_plan(conn: sqlite3.Connection, parsed_rows: list[dict]) -> dict:
    existing = existing_records_for_import(conn)
    by_order_no: dict[str, list[dict]] = defaultdict(list)
    by_client_task: dict[tuple[str, str], list[dict]] = defaultdict(list)
    by_task_name: dict[str, list[dict]] = defaultdict(list)
    by_blank_client: dict[str, list[dict]] = defaultdict(list)
    by_client_price: dict[tuple[str, float], list[dict]] = defaultdict(list)
    for row in sorted(existing, key=lambda item: item["id"]):
        wechat = clean_text(row["wechat"])
        task_name = clean_text(row["task_name"])
        order_no = clean_text(row["order_no"])
        if order_no:
            by_order_no[order_no].append(row)
        if wechat and task_name:
            by_client_task[(wechat, task_name)].append(row)
        if task_name:
            by_task_name[task_name].append(row)
        if wechat and not task_name and not order_no:
            by_blank_client[wechat].append(row)
        if wechat:
            by_client_price[(wechat, as_float(row["price"]))].append(row)
    consumed_ids: set[int] = set()

    def take_candidate(candidates: list[dict]) -> dict | None:
        for candidate in candidates:
            if candidate["id"] not in consumed_ids:
                consumed_ids.add(candidate["id"])
                return candidate
        return None

    items = []
    totals = {
        "parsed_count": len(parsed_rows),
        "created_count": 0,
        "updated_count": 0,
        "unchanged_count": 0,
        "skipped_count": 0,
        "payment_adjusted_count": 0,
        "import_price": round(sum(row["price"] for row in parsed_rows), 2),
        "import_paid": round(sum(row["paid"] for row in parsed_rows), 2),
        "import_remaining": round(sum(row["price"] - row["paid"] for row in parsed_rows), 2),
    }
    for row in parsed_rows:
        match = take_candidate(by_order_no.get(row["order_no"], [])) if row["order_no"] else None
        match_type = "单号" if match else ""
        if not match and row["wechat"] and row["task_name"]:
            match = take_candidate(by_client_task.get((row["wechat"], row["task_name"]), []))
            match_type = "客户+任务" if match else ""
        if not match and row["task_name"]:
            match = take_candidate(by_task_name.get(row["task_name"], []))
            match_type = "任务名" if match else ""
        if not match and row["wechat"] and not row["task_name"] and not row["order_no"]:
            match = take_candidate(by_blank_client.get(row["wechat"], []))
            match_type = "客户空白行顺序" if match else ""
        if not match and row["wechat"]:
            match = take_candidate(by_client_price.get((row["wechat"], row["price"]), []))
            match_type = "客户+金额" if match else ""
        if not match:
            totals["created_count"] += 1
            if row["paid"]:
                totals["payment_adjusted_count"] += 1
            items.append(
                {
                    "action": "create",
                    "record_id": None,
                    "match_type": "",
                    "payment_delta": row["paid"],
                    "changed_fields": ["new"],
                    "row": row,
                }
            )
            continue
        changed_fields = diff_order_fields(match, row)
        payment_delta = round(row["paid"] - as_float(match.get("paid")), 2)
        if abs(payment_delta) >= 0.01:
            totals["payment_adjusted_count"] += 1
        action = "update" if changed_fields or abs(payment_delta) >= 0.01 else "unchanged"
        if action == "update":
            totals["updated_count"] += 1
        else:
            totals["unchanged_count"] += 1
        items.append(
            {
                "action": action,
                "record_id": match["id"],
                "match_type": match_type,
                "payment_delta": payment_delta,
                "changed_fields": changed_fields,
                "row": row,
                "existing": {
                    "wechat": match["wechat"] or "",
                    "task_name": match["task_name"] or "",
                    "order_no": match["order_no"] or "",
                    "deadline_status": match["deadline_status"] or "",
                    "accepted_date": match["accepted_date"] or "",
                    "price": as_float(match["price"]),
                    "paid": as_float(match["paid"]),
                },
            }
        )
    return {**totals, "items": items}


def import_rows_to_db(parsed: dict, source_file: str) -> dict:
    with get_conn() as conn:
        plan = build_import_plan(conn, parsed["rows"])
        version_id = create_history_version(
            conn,
            f"导入前快照 {now_text()}",
            reason="Excel 导入前自动保存",
            source_file=source_file,
            summary={key: value for key, value in plan.items() if key != "items"},
        )
        for item in plan["items"]:
            row = item["row"]
            accepted_date = row["accepted_date"] or today_text()
            if item["action"] == "create":
                cur = conn.execute(
                    """
                    INSERT INTO records
                    (wechat, task_name, order_no, deadline_status, accepted_date, price,
                     paid, remaining, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        row["wechat"], row["task_name"], row["order_no"], row["deadline_status"],
                        accepted_date, row["price"], 0, row["price"], now_text(), now_text(),
                    ),
                )
                record_id = cur.lastrowid
                if abs(row["paid"]) >= 0.01:
                    conn.execute(
                        """
                        INSERT INTO payments (record_id, pay_date, amount, note, created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?)
                        """,
                        (record_id, accepted_date, row["paid"], f"Excel导入初始到账：{source_file}", now_text(), now_text()),
                    )
            elif item["action"] == "update":
                existing = item.get("existing", {})
                merged = {
                    "wechat": row["wechat"] or existing.get("wechat", ""),
                    "task_name": row["task_name"] or existing.get("task_name", ""),
                    "order_no": row["order_no"] or existing.get("order_no", ""),
                    "deadline_status": row["deadline_status"] or existing.get("deadline_status", ""),
                    "accepted_date": row["accepted_date"] or existing.get("accepted_date", ""),
                    "price": row["price"],
                }
                update_values = {
                    "wechat": merged["wechat"],
                    "task_name": merged["task_name"],
                    "order_no": merged["order_no"],
                    "deadline_status": merged["deadline_status"],
                    "price": merged["price"],
                    "updated_at": now_text(),
                    "id": item["record_id"],
                }
                if merged["accepted_date"]:
                    conn.execute(
                        """
                        UPDATE records
                        SET wechat = ?, task_name = ?, order_no = ?, deadline_status = ?,
                            accepted_date = ?, price = ?, updated_at = ?
                        WHERE id = ?
                        """,
                        (
                            update_values["wechat"], update_values["task_name"], update_values["order_no"],
                            update_values["deadline_status"], merged["accepted_date"], update_values["price"],
                            update_values["updated_at"], update_values["id"],
                        ),
                    )
                else:
                    conn.execute(
                        """
                        UPDATE records
                        SET wechat = ?, task_name = ?, order_no = ?, deadline_status = ?,
                            price = ?, updated_at = ?
                        WHERE id = ?
                        """,
                        (
                            update_values["wechat"], update_values["task_name"], update_values["order_no"],
                            update_values["deadline_status"], update_values["price"],
                            update_values["updated_at"], update_values["id"],
                        ),
                    )
                if abs(item["payment_delta"]) >= 0.01:
                    conn.execute(
                        """
                        INSERT INTO payments (record_id, pay_date, amount, note, created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?)
                        """,
                        (
                            item["record_id"], today_text(), item["payment_delta"],
                            f"Excel导入到账差额校正：表格到账 {row['paid']}，系统原到账 {item['existing']['paid']}",
                            now_text(), now_text(),
                        ),
                    )
        summary = {key: value for key, value in plan.items() if key != "items"}
        summary["version_id"] = version_id
        summary["preview"] = plan["items"][:20]
        conn.execute(
            """
            INSERT INTO import_batches
            (version_id, imported_at, source_file, sheet_name, parsed_count, created_count,
             updated_count, unchanged_count, skipped_count, payment_adjusted_count, summary_json)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                version_id, now_text(), source_file, parsed["sheet_name"], summary["parsed_count"],
                summary["created_count"], summary["updated_count"], summary["unchanged_count"],
                summary["skipped_count"], summary["payment_adjusted_count"],
                json.dumps(summary, ensure_ascii=False),
            ),
        )
        conn.commit()
    return summary


def month_key(day: str) -> str:
    return clean_text(day)[:7] or today_text()[:7]


def series_from_map(values: dict[str, float], limit: int | None = None) -> list[dict]:
    items = [{"label": k, "value": round(v, 2)} for k, v in sorted(values.items())]
    if limit and len(items) > limit:
        items = items[-limit:]
    return items


def build_stats(records: list[dict], payments: list[dict] | None = None, expenses: list[dict] | None = None) -> dict:
    payments = payments if payments is not None else fetch_payments()
    expenses = expenses if expenses is not None else fetch_expenses()

    total_price = sum(r["total_price"] for r in records)
    total_adjustment = sum(r["adjustment"] for r in records)
    total_paid = sum(p["amount"] for p in payments)
    total_remaining = total_price - total_paid
    total_expense = sum(e["amount"] for e in expenses)
    today = today_text()
    this_month = today[:7]
    today_date = date.today()
    unpaid_count = 0
    overdue_count = 0
    due_soon_count = 0

    status: dict[str, int] = {}
    clients: dict[str, dict] = {}
    for r in records:
        bucket = status_bucket(r["deadline_status"])
        status[bucket] = status.get(bucket, 0) + 1
        if r["remaining"] > 0:
            unpaid_count += 1
            try:
                deadline_day = datetime.strptime(clean_text(r["deadline_status"])[:10], "%Y-%m-%d").date()
                delta_days = (deadline_day - today_date).days
                if delta_days < 0:
                    overdue_count += 1
                elif delta_days <= 7:
                    due_soon_count += 1
            except ValueError:
                pass
        client = r["wechat"] or "未填写"
        clients.setdefault(client, {"client": client, "count": 0, "price": 0, "paid": 0, "remaining": 0})
        clients[client]["count"] += 1
        clients[client]["price"] += r["total_price"]
        clients[client]["paid"] += r["paid"]
        clients[client]["remaining"] += r["remaining"]

    daily_income = defaultdict(float)
    monthly_income = defaultdict(float)
    for p in payments:
        daily_income[p["pay_date"]] += p["amount"]
        monthly_income[month_key(p["pay_date"])] += p["amount"]

    daily_expense = defaultdict(float)
    monthly_expense = defaultdict(float)
    for e in expenses:
        daily_expense[e["expense_date"]] += e["amount"]
        monthly_expense[month_key(e["expense_date"])] += e["amount"]

    daily_cashflow = {}
    for k in sorted(set(daily_income) | set(daily_expense)):
        daily_cashflow[k] = round(daily_income.get(k, 0) - daily_expense.get(k, 0), 2)
    monthly_cashflow = {}
    for k in sorted(set(monthly_income) | set(monthly_expense)):
        monthly_cashflow[k] = round(monthly_income.get(k, 0) - monthly_expense.get(k, 0), 2)

    top_clients = sorted(clients.values(), key=lambda x: x["price"], reverse=True)[:10]
    for item in top_clients:
        for key in ("price", "paid", "remaining"):
            item[key] = round(item[key], 2)

    return {
        "summary": {
            "count": len(records),
            "total_price": round(total_price, 2),
            "total_adjustment": round(total_adjustment, 2),
            "total_paid": round(total_paid, 2),
            "total_remaining": round(total_remaining, 2),
            "paid_rate": round((total_paid / total_price * 100) if total_price else 0, 2),
            "total_expense": round(total_expense, 2),
            "net_income": round(total_paid - total_expense, 2),
            "today_income": round(daily_income.get(today, 0), 2),
            "month_income": round(monthly_income.get(this_month, 0), 2),
            "today_expense": round(daily_expense.get(today, 0), 2),
            "month_expense": round(monthly_expense.get(this_month, 0), 2),
            "today_net": round(daily_income.get(today, 0) - daily_expense.get(today, 0), 2),
            "month_net": round(monthly_income.get(this_month, 0) - monthly_expense.get(this_month, 0), 2),
            "avg_order_value": round((total_price / len(records)) if records else 0, 2),
            "unpaid_count": unpaid_count,
            "overdue_count": overdue_count,
            "due_soon_count": due_soon_count,
        },
        "status": [{"name": k, "value": v} for k, v in sorted(status.items())],
        "top_clients": top_clients,
        "daily_income": series_from_map(daily_income, 45),
        "monthly_income": series_from_map(monthly_income, 18),
        "daily_expense": series_from_map(daily_expense, 45),
        "monthly_expense": series_from_map(monthly_expense, 18),
        "daily_cashflow": series_from_map(daily_cashflow, 45),
        "monthly_cashflow": series_from_map(monthly_cashflow, 18),
    }


@app.route("/health")
def health():
    return "OK：项目收款统计系统正在运行", 200


@app.route("/")
def root():
    return redirect("/dashboard")


@app.route("/pwa")
@app.route("/pwa/")
def pwa_index():
    return app.send_static_file("pwa/index.html")


@app.route("/sync")
def sync_status_page():
    return render_template("sync.html", page="sync")


@app.route("/dashboard")
def dashboard_page():
    return render_template("dashboard.html", page="dashboard")


@app.route("/orders")
def orders_page():
    return render_template("orders.html", page="orders")


@app.route("/payments")
def payments_page():
    return render_template("payments.html", page="payments")


@app.route("/expenses")
def expenses_page():
    return render_template("expenses.html", page="expenses")


@app.route("/import-export")
def import_export_page():
    return render_template("import-export.html", page="import-export")


@app.errorhandler(404)
def page_not_found(error):
    if request.path.startswith("/api/") or request.path.startswith("/export/"):
        return jsonify({"ok": False, "error": "接口不存在", "path": request.path}), 404
    return redirect("/dashboard")


@app.route("/api/records", methods=["GET"])
def api_records():
    return jsonify(fetch_records())


@app.route("/api/stats", methods=["GET"])
def api_stats():
    return jsonify(build_stats(fetch_records()))


@app.route("/api/history", methods=["GET"])
def api_history_versions():
    limit = int(request.args.get("limit", 30))
    return jsonify(fetch_history_versions(limit))


@app.route("/api/history/<int:version_id>/rollback", methods=["POST"])
def api_rollback(version_id: int):
    try:
        backup_id = restore_history_version(version_id)
    except ValueError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 404
    except Exception as exc:
        return jsonify({"ok": False, "error": f"回退失败：{exc}"}), 500
    return jsonify({"ok": True, "backup_id": backup_id})


def uploaded_excel_file():
    file = request.files.get("file")
    if not file or not file.filename:
        raise ValueError("请选择要导入的 Excel 文件")
    ext = Path(file.filename).suffix.lower()
    if ext not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        raise ValueError("请上传 .xlsx 或 .xlsm 格式的表格")
    return file


@app.route("/api/import/preview", methods=["POST"])
def api_import_preview():
    try:
        file = uploaded_excel_file()
        parsed = parse_import_workbook(file)
        with get_conn() as conn:
            plan = build_import_plan(conn, parsed["rows"])
        summary = {key: value for key, value in plan.items() if key != "items"}
        return jsonify(
            {
                "ok": True,
                "filename": file.filename,
                "sheet_name": parsed["sheet_name"],
                "header_row": parsed["header_row"],
                "columns": parsed["columns"],
                "summary": summary,
                "skipped": parsed["skipped"][:30],
                "preview": plan["items"][:40],
            }
        )
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400


@app.route("/api/import/xlsx", methods=["POST"])
def api_import_xlsx():
    try:
        file = uploaded_excel_file()
        parsed = parse_import_workbook(file)
        summary = import_rows_to_db(parsed, file.filename)
        return jsonify({"ok": True, "filename": file.filename, "sheet_name": parsed["sheet_name"], "summary": summary})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400


@app.route("/api/records", methods=["POST"])
def api_create_record():
    payload = request.get_json(force=True) or {}
    data = normalize_order(payload)
    initial_paid = as_float(payload.get("initial_paid"))
    initial_pay_date = clean_date(payload.get("initial_pay_date") or data["accepted_date"])
    with get_conn() as conn:
        cur = conn.execute(
            """
            INSERT INTO records
            (wechat, task_name, order_no, deadline_status, accepted_date, price, paid, remaining, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                data["wechat"], data["task_name"], data["order_no"], data["deadline_status"],
                data["accepted_date"], data["price"], 0, data["price"], now_text(), now_text(),
            ),
        )
        record_id = cur.lastrowid
        log_change(conn, "records", record_id, "INSERT", serialize_record(conn, "records", record_id))
        if initial_paid:
            pay_cur = conn.execute(
                """
                INSERT INTO payments (record_id, pay_date, amount, note, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (record_id, initial_pay_date, initial_paid, "新增订单首笔付款", now_text(), now_text()),
            )
            log_change(conn, "payments", pay_cur.lastrowid, "INSERT", serialize_record(conn, "payments", pay_cur.lastrowid))
        conn.commit()
    return jsonify({"ok": True, "id": record_id})


@app.route("/api/records/<int:record_id>", methods=["PUT"])
def api_update_record(record_id: int):
    data = normalize_order(request.get_json(force=True) or {})
    with get_conn() as conn:
        conn.execute(
            """
            UPDATE records
            SET wechat = ?, task_name = ?, order_no = ?, deadline_status = ?,
                accepted_date = ?, price = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                data["wechat"], data["task_name"], data["order_no"], data["deadline_status"],
                data["accepted_date"], data["price"], now_text(), record_id,
            ),
        )
        log_change(conn, "records", record_id, "UPDATE", serialize_record(conn, "records", record_id))
        conn.commit()
    return jsonify({"ok": True})


@app.route("/api/records/<int:record_id>", methods=["DELETE"])
def api_delete_record(record_id: int):
    with get_conn() as conn:
        # 级联删除前记录关联数据
        conn.execute("DELETE FROM payments WHERE record_id = ?", (record_id,))
        conn.execute("DELETE FROM budget_changes WHERE record_id = ?", (record_id,))
        conn.execute("DELETE FROM records WHERE id = ?", (record_id,))
        log_change(conn, "records", record_id, "DELETE", {"id": record_id})
        conn.commit()
    return jsonify({"ok": True})


@app.route("/api/records/<int:record_id>/payments", methods=["GET"])
def api_record_payments(record_id: int):
    return jsonify(fetch_payments(record_id))


@app.route("/api/records/<int:record_id>/payments", methods=["POST"])
def api_create_payment(record_id: int):
    data = normalize_payment(request.get_json(force=True) or {})
    with get_conn() as conn:
        cur = conn.execute(
            """
            INSERT INTO payments (record_id, pay_date, amount, note, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (record_id, data["pay_date"], data["amount"], data["note"], now_text(), now_text()),
        )
        pid = cur.lastrowid
        log_change(conn, "payments", pid, "INSERT", serialize_record(conn, "payments", pid))
        conn.commit()
    return jsonify({"ok": True, "id": pid})


@app.route("/api/payments/<int:payment_id>", methods=["PUT"])
def api_update_payment(payment_id: int):
    data = normalize_payment(request.get_json(force=True) or {})
    with get_conn() as conn:
        conn.execute(
            """
            UPDATE payments
            SET pay_date = ?, amount = ?, note = ?, updated_at = ?
            WHERE id = ?
            """,
            (data["pay_date"], data["amount"], data["note"], now_text(), payment_id),
        )
        log_change(conn, "payments", payment_id, "UPDATE", serialize_record(conn, "payments", payment_id))
        conn.commit()
    return jsonify({"ok": True})


@app.route("/api/payments/<int:payment_id>", methods=["DELETE"])
def api_delete_payment(payment_id: int):
    with get_conn() as conn:
        conn.execute("DELETE FROM payments WHERE id = ?", (payment_id,))
        log_change(conn, "payments", payment_id, "DELETE", {"id": payment_id})
        conn.commit()
    return jsonify({"ok": True})


@app.route("/api/records/<int:record_id>/adjustments", methods=["GET"])
def api_record_adjustments(record_id: int):
    return jsonify(fetch_adjustments(record_id))


@app.route("/api/records/<int:record_id>/adjustments", methods=["POST"])
def api_create_adjustment(record_id: int):
    data = normalize_adjustment(request.get_json(force=True) or {})
    with get_conn() as conn:
        cur = conn.execute(
            """
            INSERT INTO budget_changes (record_id, change_date, amount, note, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (record_id, data["change_date"], data["amount"], data["note"], now_text(), now_text()),
        )
        aid = cur.lastrowid
        log_change(conn, "budget_changes", aid, "INSERT", serialize_record(conn, "budget_changes", aid))
        conn.commit()
    return jsonify({"ok": True, "id": aid})


@app.route("/api/adjustments/<int:adjustment_id>", methods=["PUT"])
def api_update_adjustment(adjustment_id: int):
    data = normalize_adjustment(request.get_json(force=True) or {})
    with get_conn() as conn:
        conn.execute(
            """
            UPDATE budget_changes
            SET change_date = ?, amount = ?, note = ?, updated_at = ?
            WHERE id = ?
            """,
            (data["change_date"], data["amount"], data["note"], now_text(), adjustment_id),
        )
        log_change(conn, "budget_changes", adjustment_id, "UPDATE", serialize_record(conn, "budget_changes", adjustment_id))
        conn.commit()
    return jsonify({"ok": True})


@app.route("/api/adjustments/<int:adjustment_id>", methods=["DELETE"])
def api_delete_adjustment(adjustment_id: int):
    with get_conn() as conn:
        conn.execute("DELETE FROM budget_changes WHERE id = ?", (adjustment_id,))
        log_change(conn, "budget_changes", adjustment_id, "DELETE", {"id": adjustment_id})
        conn.commit()
    return jsonify({"ok": True})


@app.route("/api/expenses", methods=["GET"])
def api_expenses():
    return jsonify(fetch_expenses())


@app.route("/api/expenses", methods=["POST"])
def api_create_expense():
    data = normalize_expense(request.get_json(force=True) or {})
    with get_conn() as conn:
        cur = conn.execute(
            """
            INSERT INTO expenses (expense_date, name, amount, note, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (data["expense_date"], data["name"], data["amount"], data["note"], now_text(), now_text()),
        )
        eid = cur.lastrowid
        log_change(conn, "expenses", eid, "INSERT", serialize_record(conn, "expenses", eid))
        conn.commit()
    return jsonify({"ok": True, "id": eid})


@app.route("/api/expenses/<int:expense_id>", methods=["PUT"])
def api_update_expense(expense_id: int):
    data = normalize_expense(request.get_json(force=True) or {})
    with get_conn() as conn:
        conn.execute(
            """
            UPDATE expenses
            SET expense_date = ?, name = ?, amount = ?, note = ?, updated_at = ?
            WHERE id = ?
            """,
            (data["expense_date"], data["name"], data["amount"], data["note"], now_text(), expense_id),
        )
        log_change(conn, "expenses", expense_id, "UPDATE", serialize_record(conn, "expenses", expense_id))
        conn.commit()
    return jsonify({"ok": True})


@app.route("/api/expenses/<int:expense_id>", methods=["DELETE"])
def api_delete_expense(expense_id: int):
    with get_conn() as conn:
        conn.execute("DELETE FROM expenses WHERE id = ?", (expense_id,))
        log_change(conn, "expenses", expense_id, "DELETE", {"id": expense_id})
        conn.commit()
    return jsonify({"ok": True})


def style_range(ws, cell_range, fill=None, font=None, border=None, alignment=None):
    for row in ws[cell_range]:
        for cell in row:
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            if border:
                cell.border = border
            if alignment:
                cell.alignment = alignment


# ==================== Sync Endpoints ====================

@app.route("/api/sync/info", methods=["GET"])
def api_sync_info():
    with get_conn() as conn:
        record_count = conn.execute("SELECT COUNT(*) FROM records").fetchone()[0]
        payment_count = conn.execute("SELECT COUNT(*) FROM payments").fetchone()[0]
        adjustment_count = conn.execute("SELECT COUNT(*) FROM budget_changes").fetchone()[0]
        expense_count = conn.execute("SELECT COUNT(*) FROM expenses").fetchone()[0]
        last_updated = conn.execute("SELECT MAX(updated_at) FROM records").fetchone()[0] or ""
        total_paid = conn.execute("SELECT COALESCE(SUM(amount), 0) FROM payments").fetchone()[0]
        total_remaining = conn.execute(
            """
            SELECT COALESCE(SUM(r.price + COALESCE(b.adj, 0) - COALESCE(p.paid, 0)), 0)
            FROM records r
            LEFT JOIN (SELECT record_id, SUM(amount) AS adj FROM budget_changes GROUP BY record_id) b ON b.record_id = r.id
            LEFT JOIN (SELECT record_id, SUM(amount) AS paid FROM payments GROUP BY record_id) p ON p.record_id = r.id
            WHERE r.price + COALESCE(b.adj, 0) - COALESCE(p.paid, 0) > 0
            """
        ).fetchone()[0]
    return jsonify({
        "ok": True,
        "info": {
            "record_count": record_count,
            "payment_count": payment_count,
            "adjustment_count": adjustment_count,
            "expense_count": expense_count,
            "last_updated": last_updated,
            "total_paid": round(as_float(total_paid), 2),
            "total_remaining": round(as_float(total_remaining), 2),
            "db_size_kb": round(DB_PATH.stat().st_size / 1024, 1) if DB_PATH.exists() else 0,
        }
    })


@app.route("/api/sync/download", methods=["GET"])
def api_sync_download():
    if not DB_PATH.exists():
        return jsonify({"ok": False, "error": "数据库不存在"}), 404
    return send_file(
        DB_PATH,
        as_attachment=True,
        download_name="records.db",
        mimetype="application/octet-stream",
    )


@app.route("/api/sync/upload", methods=["POST"])
def api_sync_upload():
    file = request.files.get("db")
    if not file or not file.filename:
        return jsonify({"ok": False, "error": "请选择数据库文件"}), 400

    tmp_path = DATA_DIR / "_sync_temp.db"
    file.save(tmp_path)
    try:
        test_conn = sqlite3.connect(tmp_path)
        tables = test_conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()
        table_names = {t[0] for t in tables}
        required = {"records", "payments", "budget_changes", "expenses"}
        if not required.issubset(table_names):
            test_conn.close()
            tmp_path.unlink(missing_ok=True)
            return jsonify({"ok": False, "error": f"文件不是有效的数据库，缺少表：{required - table_names}"}), 400
        test_conn.close()
    except Exception as exc:
        tmp_path.unlink(missing_ok=True)
        return jsonify({"ok": False, "error": f"无法读取数据库文件：{exc}"}), 400

    with get_conn() as conn:
        backup_id = create_history_version(
            conn,
            f"同步前自动备份 {now_text()}",
            reason="接受同步数据库前自动保存",
            source_file="sync_upload",
        )
        conn.commit()

    tmp_path.replace(DB_PATH)
    return jsonify({"ok": True, "backup_id": backup_id, "message": "数据库已同步，旧数据已备份"})


# ==================== P2P Sync API ====================

@app.route("/api/sync/device", methods=["GET"])
def api_sync_device():
    """返回本设备标识和同步状态。"""
    state = get_sync_state()
    with get_conn() as conn:
        pending = conn.execute("SELECT COUNT(*) FROM change_log WHERE synced = 0").fetchone()[0]
        total = conn.execute("SELECT COUNT(*) FROM change_log").fetchone()[0]
    return jsonify({
        "ok": True,
        "device_id": state["device_id"],
        "display_name": state["display_name"],
        "last_pull_seq": state["last_pull_seq"],
        "last_push_at": state["last_push_at"],
        "pending_changes": pending,
        "total_changes": total,
        "relay_url": os.environ.get("SYNC_RELAY_URL", "https://income-sync-relay.onrender.com"),
        "sync_interval_sec": int(os.environ.get("SYNC_INTERVAL_SEC", 30)),
    })


@app.route("/api/sync/push", methods=["POST"])
def api_sync_push():
    """手动推送本地变更。"""
    with get_conn() as conn:
        result = sync_now(conn)
    return jsonify(result)


@app.route("/api/sync/state", methods=["GET"])
def api_sync_state():
    """返回当前同步引擎的完整状态。"""
    state = get_sync_state()
    with get_conn() as conn:
        pending = conn.execute("SELECT COUNT(*) FROM change_log WHERE synced = 0").fetchone()[0]
        total = conn.execute("SELECT COUNT(*) FROM change_log").fetchone()[0]
        recent = conn.execute(
            "SELECT * FROM change_log ORDER BY id DESC LIMIT 20"
        ).fetchall()
    return jsonify({
        "ok": True,
        "device": {
            "device_id": state["device_id"],
            "display_name": state["display_name"],
        },
        "sync": {
            "last_pull_seq": state["last_pull_seq"],
            "last_push_at": state["last_push_at"],
            "pending_changes": pending,
            "total_changes": total,
        },
        "recent_changes": [
            {
                "change_id": r["change_id"],
                "table_name": r["table_name"],
                "row_id": r["row_id"],
                "operation": r["operation"],
                "synced": bool(r["synced"]),
            }
            for r in recent
        ],
        "relay_url": os.environ.get("SYNC_RELAY_URL", "https://income-sync-relay.onrender.com"),
    })


@app.route("/export/xlsx", methods=["GET"])
def export_xlsx():
    try:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        return jsonify({"ok": False, "error": f"缺少 openpyxl：{exc}，请先 pip install -r requirements.txt"}), 500

    records = fetch_records()
    payments = fetch_payments()
    adjustments = fetch_adjustments()
    expenses = fetch_expenses()
    stats = build_stats(records, payments, expenses)

    wb = Workbook()
    ws = wb.active
    ws.title = "订单明细"
    pay_ws = wb.create_sheet("客户付款流水")
    adj_ws = wb.create_sheet("预算增加流水")
    exp_ws = wb.create_sheet("工具费用支出")
    dash = wb.create_sheet("收支看板")
    guide = wb.create_sheet("使用说明")

    dark = "0F172A"
    blue = "1D4ED8"
    light_blue = "DBEAFE"
    green = "DCFCE7"
    red = "FEE2E2"
    purple = "F3E8FF"
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor=blue)
    header_font = Font(color="FFFFFF", bold=True)
    title_fill = PatternFill("solid", fgColor=dark)
    title_font = Font(color="FFFFFF", bold=True, size=18)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    order_headers = ["ID", "微信号", "任务名字", "单号", "接单日期", "截止/状态", "基础价格", "增加预算", "订单总价", "客户已付", "剩余未收"]
    ws.append(order_headers)
    for r_idx, r in enumerate(records, start=2):
        ws.append([r["id"], r["wechat"], r["task_name"], r["order_no"], r["accepted_date"], r["deadline_status"], r["price"], None, None, None, None])
        ws[f"H{r_idx}"] = f"=SUMIFS('预算增加流水'!D:D,'预算增加流水'!B:B,A{r_idx})"
        ws[f"I{r_idx}"] = f"=G{r_idx}+H{r_idx}"
        ws[f"J{r_idx}"] = f"=SUMIFS('客户付款流水'!D:D,'客户付款流水'!B:B,A{r_idx})"
        ws[f"K{r_idx}"] = f"=I{r_idx}-J{r_idx}"
    total_row = len(records) + 2
    ws[f"A{total_row}"] = "合计"
    for col in range(7, 12):
        letter = get_column_letter(col)
        ws[f"{letter}{total_row}"] = f"=SUM({letter}2:{letter}{total_row-1})"

    pay_ws.append(["ID", "订单ID", "付款日期", "金额", "备注", "微信号", "任务名字", "单号"])
    for p in payments:
        pay_ws.append([p["id"], p["record_id"], p["pay_date"], p["amount"], p["note"], p["wechat"], p["task_name"], p["order_no"]])

    adj_ws.append(["ID", "订单ID", "增加日期", "金额", "说明", "微信号", "任务名字", "单号"])
    for a in adjustments:
        adj_ws.append([a["id"], a["record_id"], a["change_date"], a["amount"], a["note"], a["wechat"], a["task_name"], a["order_no"]])

    exp_ws.append(["ID", "费用日期", "工具/用途", "金额", "备注"])
    for e in expenses:
        exp_ws.append([e["id"], e["expense_date"], e["name"], e["amount"], e["note"]])

    for sheet in [ws, pay_ws, adj_ws, exp_ws]:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.freeze_panes = "A2"

    for row in ws.iter_rows(min_row=2, max_row=total_row, min_col=7, max_col=11):
        for cell in row:
            cell.number_format = '¥#,##0.00'
    for row in pay_ws.iter_rows(min_row=2, max_row=pay_ws.max_row, min_col=4, max_col=4):
        row[0].number_format = '¥#,##0.00'
    for row in adj_ws.iter_rows(min_row=2, max_row=adj_ws.max_row, min_col=4, max_col=4):
        row[0].number_format = '¥#,##0.00'
    for row in exp_ws.iter_rows(min_row=2, max_row=exp_ws.max_row, min_col=4, max_col=4):
        row[0].number_format = '¥#,##0.00'

    for cell in ws[total_row]:
        cell.fill = PatternFill("solid", fgColor=light_blue)
        cell.font = Font(bold=True)

    widths = {
        ws: [8, 18, 38, 18, 14, 16, 14, 14, 14, 14, 14],
        pay_ws: [8, 10, 14, 14, 26, 18, 34, 18],
        adj_ws: [8, 10, 14, 14, 26, 18, 34, 18],
        exp_ws: [8, 14, 28, 14, 36],
    }
    for sheet, sheet_widths in widths.items():
        for idx, width in enumerate(sheet_widths, start=1):
            sheet.column_dimensions[get_column_letter(idx)].width = width

    # Dashboard
    dash.merge_cells("A1:L1")
    dash["A1"] = "项目收支数据看板"
    dash["A1"].fill = title_fill
    dash["A1"].font = title_font
    dash["A1"].alignment = center

    kpis = [
        ("订单数", stats["summary"]["count"], "A3", light_blue),
        ("订单总价", stats["summary"]["total_price"], "C3", light_blue),
        ("客户已付", stats["summary"]["total_paid"], "E3", green),
        ("剩余未收", stats["summary"]["total_remaining"], "G3", red),
        ("总支出", stats["summary"]["total_expense"], "A6", purple),
        ("净收入", stats["summary"]["net_income"], "C6", green),
        ("今日收入", stats["summary"]["today_income"], "E6", green),
        ("本月收入", stats["summary"]["month_income"], "G6", green),
        ("今日支出", stats["summary"]["today_expense"], "I3", red),
        ("本月支出", stats["summary"]["month_expense"], "K3", red),
        ("今日净额", stats["summary"]["today_net"], "I6", purple),
        ("本月净额", stats["summary"]["month_net"], "K6", purple),
    ]
    for label, value, anchor, fill_color in kpis:
        col = anchor[0]
        row = int(anchor[1:])
        dash[f"{col}{row}"] = label
        dash[f"{col}{row+1}"] = value
        dash[f"{col}{row}"].font = Font(bold=True, color="334155")
        dash[f"{col}{row+1}"].font = Font(bold=True, color="0F172A", size=14)
        dash[f"{col}{row}"].fill = PatternFill("solid", fgColor=fill_color)
        dash[f"{col}{row+1}"].fill = PatternFill("solid", fgColor=fill_color)
        dash[f"{col}{row}"].border = border
        dash[f"{col}{row+1}"].border = border
        dash[f"{col}{row}"].alignment = center
        dash[f"{col}{row+1}"].alignment = center
        if label != "订单数":
            dash[f"{col}{row+1}"].number_format = '¥#,##0.00'

    # 每日收支表：日期 / 收入 / 支出 / 净额，便于做多曲线图。
    daily_labels = sorted({x["label"] for x in stats["daily_income"]} | {x["label"] for x in stats["daily_expense"]} | {x["label"] for x in stats["daily_cashflow"]})
    daily_income_map = {x["label"]: x["value"] for x in stats["daily_income"]}
    daily_expense_map = {x["label"]: x["value"] for x in stats["daily_expense"]}
    daily_cashflow_map = {x["label"]: x["value"] for x in stats["daily_cashflow"]}
    if not daily_labels:
        daily_labels = [today_text()]
    daily_start = 10
    for col, header in enumerate(["日期", "收入", "支出", "净额"], start=1):
        cell = dash.cell(daily_start, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
    for i, label in enumerate(daily_labels[-45:], start=daily_start + 1):
        dash.cell(i, 1, label)
        dash.cell(i, 2, daily_income_map.get(label, 0))
        dash.cell(i, 3, daily_expense_map.get(label, 0))
        dash.cell(i, 4, daily_cashflow_map.get(label, 0))
        for col in range(2, 5):
            dash.cell(i, col).number_format = '¥#,##0.00'

    # 每月收支表：月份 / 收入 / 支出 / 净额。
    monthly_labels = sorted({x["label"] for x in stats["monthly_income"]} | {x["label"] for x in stats["monthly_expense"]} | {x["label"] for x in stats["monthly_cashflow"]})
    monthly_income_map = {x["label"]: x["value"] for x in stats["monthly_income"]}
    monthly_expense_map = {x["label"]: x["value"] for x in stats["monthly_expense"]}
    monthly_cashflow_map = {x["label"]: x["value"] for x in stats["monthly_cashflow"]}
    if not monthly_labels:
        monthly_labels = [today_text()[:7]]
    monthly_start = 10
    for col, header in enumerate(["月份", "收入", "支出", "净额"], start=6):
        cell = dash.cell(monthly_start, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
    for i, label in enumerate(monthly_labels[-18:], start=monthly_start + 1):
        dash.cell(i, 6, label)
        dash.cell(i, 7, monthly_income_map.get(label, 0))
        dash.cell(i, 8, monthly_expense_map.get(label, 0))
        dash.cell(i, 9, monthly_cashflow_map.get(label, 0))
        for col in range(7, 10):
            dash.cell(i, col).number_format = '¥#,##0.00'

    status_start = max(daily_start + len(daily_labels[-45:]), monthly_start + len(monthly_labels[-18:])) + 4
    dash[f"A{status_start}"] = "状态"
    dash[f"B{status_start}"] = "数量"
    for c in [f"A{status_start}", f"B{status_start}"]:
        dash[c].fill = header_fill
        dash[c].font = header_font
        dash[c].alignment = center
    for i, item in enumerate(stats["status"], start=status_start + 1):
        dash[f"A{i}"] = item["name"]
        dash[f"B{i}"] = item["value"]

    client_start = status_start
    dash[f"D{client_start}"] = "客户"
    dash[f"E{client_start}"] = "订单总价"
    dash[f"F{client_start}"] = "已付款"
    dash[f"G{client_start}"] = "剩余未收"
    for c in [f"D{client_start}", f"E{client_start}", f"F{client_start}", f"G{client_start}"]:
        dash[c].fill = header_fill
        dash[c].font = header_font
        dash[c].alignment = center
    for i, item in enumerate(stats["top_clients"], start=client_start + 1):
        dash[f"D{i}"] = item["client"]
        dash[f"E{i}"] = item["price"]
        dash[f"F{i}"] = item["paid"]
        dash[f"G{i}"] = item["remaining"]
        for col in ["E", "F", "G"]:
            dash[f"{col}{i}"].number_format = '¥#,##0.00'

    for row in dash.iter_rows(min_row=1, max_row=max(client_start + len(stats["top_clients"]) + 2, status_start + len(stats["status"]) + 2), max_col=12):
        for cell in row:
            if cell.value is not None:
                cell.border = border
                if cell.alignment is None:
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
    for col in range(1, 13):
        dash.column_dimensions[get_column_letter(col)].width = 15

    # Charts
    chart = LineChart()
    chart.title = "每日收入 / 支出 / 净额"
    chart.y_axis.title = "金额"
    chart.x_axis.title = "日期"
    data = Reference(dash, min_col=2, max_col=4, min_row=daily_start, max_row=daily_start + len(daily_labels[-45:]))
    cats = Reference(dash, min_col=1, min_row=daily_start + 1, max_row=daily_start + len(daily_labels[-45:]))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 9
    chart.width = 20
    dash.add_chart(chart, "J10")

    chart = LineChart()
    chart.title = "每月收入 / 支出 / 净额"
    chart.y_axis.title = "金额"
    chart.x_axis.title = "月份"
    data = Reference(dash, min_col=7, max_col=9, min_row=monthly_start, max_row=monthly_start + len(monthly_labels[-18:]))
    cats = Reference(dash, min_col=6, min_row=monthly_start + 1, max_row=monthly_start + len(monthly_labels[-18:]))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 9
    chart.width = 20
    dash.add_chart(chart, f"J{monthly_start + 18}")

    if stats["status"]:
        chart = DoughnutChart()
        chart.title = "订单状态分布"
        data = Reference(dash, min_col=2, min_row=status_start, max_row=status_start + len(stats["status"]))
        cats = Reference(dash, min_col=1, min_row=status_start + 1, max_row=status_start + len(stats["status"]))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 12
        dash.add_chart(chart, f"J{status_start}")

    if stats["top_clients"]:
        chart = BarChart()
        chart.title = "Top 10 客户金额"
        chart.y_axis.title = "金额"
        data = Reference(dash, min_col=5, max_col=7, min_row=client_start, max_row=client_start + len(stats["top_clients"]))
        cats = Reference(dash, min_col=4, min_row=client_start + 1, max_row=client_start + len(stats["top_clients"]))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 18
        dash.add_chart(chart, f"J{status_start + 14}")

    guide["A1"] = "使用说明"
    guide["A2"] = "1. 订单主表记录接单日期、基础价格；付款流水会自动汇总到订单已付金额。"
    guide["A3"] = "2. 客户每次付款都记录付款日期和金额，因此每日收入、每月收入都会自动变化。"
    guide["A4"] = "3. 预算增加流水用于记录客户临时追加预算，订单总价=基础价格+增加预算。"
    guide["A5"] = "4. 工具费用支出记录支出日期和金额，收支看板会统计净收入。"
    guide["A6"] = "5. 删除或修改订单、付款、预算增加、工具费用后，页面和导出表都会重新计算。"
    guide.column_dimensions["A"].width = 110

    tmp = NamedTemporaryFile(suffix=".xlsx", delete=False, dir=EXPORT_DIR)
    wb.save(tmp.name)
    tmp.close()
    return send_file(tmp.name, as_attachment=True, download_name="统计_收支系统导出.xlsx")


@app.route("/<path:any_path>")
def frontend_fallback(any_path: str):
    if any_path.startswith("api/") or any_path.startswith("export/"):
        return jsonify({"ok": False, "error": "接口不存在", "path": "/" + any_path}), 404
    return redirect("/dashboard")


def open_browser() -> None:
    webbrowser.open("http://127.0.0.1:5050/")


if __name__ == "__main__":
    import socket
    init_db()
    Timer(1.2, open_browser).start()

    local_ip = "127.0.0.1"
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        local_ip = s.getsockname()[0]
        s.close()
    except Exception:
        pass

    print("系统已启动：")
    print(f"  本地访问：http://127.0.0.1:5050")
    print(f"  局域网访问：http://{local_ip}:5050")
    print(f"  健康检查：http://127.0.0.1:5050/health")
    print(f"  PWA 移动端：http://{local_ip}:5050/pwa")
    print(f"  数据库位置：{DB_PATH}")
    print(f"  设备 ID：{get_device_id()}")
    print("关闭这个窗口即可停止系统。")

    # 启动后台自动同步
    start_sync_loop()
    print("后台同步已启动（每 30 秒同步一次）")

    app.run(host="0.0.0.0", port=5050, debug=False, use_reloader=False)
